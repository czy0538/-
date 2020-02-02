# czy0538 2020年2月2日 2020年2月2日 21:30:25
# v 1.1

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import win32com.client as win32

message = '请先确保导出.xlsm,需要处理的文件（直接从问卷星导出即可）都位于该目录下\n'
print(message)
filename = input('请输入文件名:')
filename2 = '导出.xlsm'


def change_xls():
    excel = win32.Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(os.path.abspath(filename))

    wb.SaveAs(os.path.abspath(filename) + "x", FileFormat=51)
    wb.Close()
    excel.Application.Quit()


def replace_xls2xlsm(sheetname):
    wb = load_workbook(filename)
    wb2 = load_workbook(filename2, read_only=False, keep_vba=True)

    ws = wb[sheetname]
    ws2 = wb2[sheetname]

    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)

    wb2.save(filename2)


def replace_xlsm2xls(sheetname):
    wb = load_workbook(filename2, read_only=False, keep_vba=True)
    wb2 = Workbook()

    ws = wb[sheetname]
    ws2 = wb2.active
    ws2.title = sheetname

    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)

    wb2.save(filename)


def run_excel_macro():
    xl = win32.Dispatch('Excel.Application')
    xl.Application.visible = False

    try:
        wb = xl.Workbooks.Open(os.path.abspath(filename2))
        xl.Application.run(filename2 + "!Formatting")
        wb.Save()
        wb.Close()

    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)

    xl.Application.Quit()
    del xl


def main():
    print("开始处理，请等待")
    change_xls()
    global filename
    filename = filename + 'x'

    sheetnames = [u'Sheet1']

    for sheetname in sheetnames:
        replace_xls2xlsm(sheetname)

    run_excel_macro()
    print("修改完成")
    os.remove(filename)

    for sheetname in sheetnames:
        replace_xlsm2xls(sheetname)


if __name__ == "__main__":
    main()
